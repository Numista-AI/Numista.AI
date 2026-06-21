"""
Read all PCS Folder Excel files and show their contents.
"""
import os, sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import glob

PCS_FOLDER = r"C:\Users\ericd\Documents\MyVertexProject\AJ's Coins\PCS Folder"

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed")
    sys.exit(1)

xlsx_files = glob.glob(os.path.join(PCS_FOLDER, '*.xlsx'))
print(f'Found {len(xlsx_files)} Excel files in PCS Folder:\n')

for fpath in sorted(xlsx_files):
    fname = os.path.basename(fpath)
    print(f'=== {fname} ===')
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                print(f'  Sheet "{sheet_name}": empty')
                continue
            # Print headers + first 20 data rows
            header = rows[0]
            data_rows = rows[1:]
            print(f'  Sheet "{sheet_name}": {len(data_rows)} data rows')
            print(f'  Headers: {list(header)}')
            for row in data_rows[:20]:
                # Print non-empty values
                row_str = ' | '.join(str(v) for v in row if v is not None and str(v).strip())
                if row_str:
                    print(f'    {row_str}')
        wb.close()
    except Exception as e:
        print(f'  ERROR: {e}')
    print()
